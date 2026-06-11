import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DARK_BLUE = "1E3A5F"
LIGHT_BLUE = "EBF4FF"
GREEN = "2ECC71"
RED = "E74C3C"


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)


def apply_currency(cell, value) -> None:
    cell.value = value
    cell.number_format = "#,##0.00"


def _write_sheet(ws, headers: list[str], rows: list[dict], mapping: list[tuple[str, str]]):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, key) in enumerate(mapping, start=1):
            val = row.get(key, "")
            if key == "amount" or key.endswith("_amount"):
                apply_currency(ws.cell(row=row_idx, column=col_idx), float(val or 0))
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(val or ""))
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)


def generate_monthly_excel(data: dict, business_name: str, month: str, year: int) -> bytes:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = f"{business_name} — {month} {year}"
    ws_summary["A1"].font = Font(bold=True, size=16, color=DARK_BLUE)

    total_sales = sum(float(t["amount"]) for t in data.get("sales", []))
    total_expenses = sum(float(t["amount"]) for t in data.get("expenses", []))
    net_profit = total_sales - total_expenses
    udhaar_outstanding = sum(
        float(r["total_amount"]) - float(r.get("paid_amount", 0))
        for r in data.get("udhaar_given", [])
        if r.get("status") != "paid"
    )

    rows = [
        ("Total Revenue (Sales)", total_sales),
        ("Total Expenses", total_expenses),
        ("Net Profit", net_profit),
        ("Outstanding Udhaar", udhaar_outstanding),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws_summary[f"A{i}"] = label
        apply_currency(ws_summary[f"B{i}"], value)
        if label == "Net Profit":
            ws_summary[f"B{i}"].font = Font(
                bold=True, color=GREEN if value >= 0 else RED, size=12
            )

    ws_sales = wb.create_sheet("Sales")
    _write_sheet(
        ws_sales,
        ["Date", "Description", "Customer", "Payment", "Amount"],
        data.get("sales", []),
        [
            ("d", "transaction_date"),
            ("desc", "description"),
            ("c", "customer_name"),
            ("p", "payment_method"),
            ("a", "amount"),
        ],
    )

    ws_exp = wb.create_sheet("Expenses")
    _write_sheet(
        ws_exp,
        ["Date", "Description", "Category", "Payment", "Amount"],
        data.get("expenses", []),
        [
            ("d", "transaction_date"),
            ("desc", "description"),
            ("cat", "category"),
            ("p", "payment_method"),
            ("a", "amount"),
        ],
    )

    ws_udhaar_g = wb.create_sheet("Udhaar Given")
    _write_sheet(
        ws_udhaar_g,
        ["Party", "Phone", "Total", "Paid", "Status", "Due"],
        data.get("udhaar_given", []),
        [
            ("n", "party_name"),
            ("ph", "party_phone"),
            ("t", "total_amount"),
            ("p", "paid_amount"),
            ("s", "status"),
            ("d", "due_date"),
        ],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
